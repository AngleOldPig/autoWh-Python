from openpyxl import load_workbook
import pandas as pd
import csv
import os
import openpyxl
import re
import numpy as np
np.set_printoptions(threshold=np.inf)

# 需要填写的参数：
# *** 注意 *** 不要填写后缀名，也不要填写文件名的日期部分
predictFileName = '5934_muestras_todos_los_meses'   # 要读取的预测数据文档名
actualFileName = 'Energía_y_potencia_Día'           # 要读取的实际数据文档名
actualFileMode = 2                                  # 实际数据读取模式
writeFileName = 'Comparativa_Susana'                # 需要录入的表格名
resultFileName = 'Comparativa_Susana-finish'        # 保存结果的文档名
# *** 注意 *** 不要填写后缀名，也不要填写文件名的日期部分
# 实际数据读取模式：
# 模式1. 数据存放在第8列，时间命名用yyyy_mm_dd表示
# 模式2. 数据存放在第2列，时间命名用yyyy-mm-dd表示，数据大小×1000


# 加载参数文件：
print("读取中。。。")

# 使用pd.read_csv需手动加入参数
# 国内（delimiter=";", decimal=",", thousands='.')
# 西班牙（delimiter=";"）


# 加工预测数据文档
data1 = pd.read_csv(predictFileName + '.csv', delimiter=";", encoding='utf-8')  # 在国内使用需手动切换国内格式
data1["Production"] = data1["Production"] * 1000
data1.to_excel('somethingYouNeed.xlsx', sheet_name='Prediction')

# 加工实际数据文档：

# 生成空模板csv文档，并添加年、月、日、小时、每日产电量，5列标题
actualProductionCsvName = "somethingYouNeedToo.csv"
with open(actualProductionCsvName, 'w') as csvfile:
    csv_writer = csv.writer(csvfile)
    csv_head = ["year", "month", "day", "hour", "production"]
    csv_writer.writerow(csv_head)

# 自动生成真实日期数字，区分模式1和模式2


# 模式1. 数据存放在第8列，时间命名用yyyy_mm_dd表示
if actualFileMode == 1:
    for y in range(2019, 2021):
        for m in range(1, 13):
            for d in range(1, 32):
                if m == 2 and d > 29:
                    continue
                if m == 4 or m == 6 or m == 9 or m == 11:
                    if d > 30:
                        continue
                # 使一位数字前面加0变成两位数字
                year = str(y)
                month = str(m)
                if m < 10:
                    month = '0' + month
                day = str(d)
                if d < 10:
                    day = '0' + day
                # 组合生成文件名
                data2FileName = actualFileName + '_' + year + '_' + month + '_' + day + '.csv'
                # 判断当前名称的文件是否存在
                if os.path.exists(data2FileName):
                    if os.path.getsize(data2FileName):
                        print('已读取' + data2FileName)
                    else:
                        print(data2FileName + '文件存在但为空')
                        continue
                else:
                    # print(data2FileName + '文件不存在')
                    continue
                # 用暂存变量dataTemp读取生成文档名指向的csv文件
                dataTemp = pd.read_csv(data2FileName,
                                       delimiter=";",       # 在国内使用需手动切换国内格式
                                       encoding='utf-8', header=None, skiprows=1, usecols=[0, 8])
                # 去除数据中的. 防止数据被识别为小数

                # 准备参数
                dataTempList = []
                a = 0
                h = 0
                num = [0.0, 0.0, 0.0, 0.0, 0.0]

                # 调试用的信息
                # print('表格模板：')
                # print(dataTempList)
                # print('选取项：')
                # print(dataTemp[0][3])
                # print('所有项：')
                # print(dataTemp)

                # 将文件中的实际值读取出并和平均值一起存入num数组
                for i in range(3, 96):
                    p = dataTemp[8][i]
                    num[(i+1) % 4] = float(p)                               # 读取每小时的4个数据单元格
                    if (i + 1) % 4 == 3:
                        num[4] = num[0] + num[1] + num[2] + num[3]
                        num[4] = num[4] / 4                             # 求它们的平均数
                        a = a + 1
                        dataTempList.append([y,m,d,a,num[4]])           # 将时间数据和每小时平均数添加进dataTempList列表
                dataTempList.append([y,m,d,24,0])

                # 将 列表list 转换为 DataFrame格式
                data2TempDataFrame = pd.DataFrame(dataTempList)
                # print('表格雏形：')
                # print(data2TempDataFrame)

                # 将时间信息与每小时平均数一并存入csv文件：
                data2TempDataFrame.to_csv(actualProductionCsvName, mode='a', header=False, index=None)


# 模式2. 数据存放在第2列，时间命名用yyyy-mm-dd表示，数据大小×1000
if actualFileMode == 2:
    for y in range(2019, 2021):
        for m in range(1, 13):
            for d in range(1, 32):
                if m == 2 and d > 29:
                    continue
                if m == 4 or m == 6 or m == 9 or m == 11:
                    if d > 30:
                        continue
                # 使一位数字前面加0变成两位数字
                year = str(y)
                month = str(m)
                if m < 10:
                    month = '0' + month
                day = str(d)
                if d < 10:
                    day = '0' + day
                # 组合生成文件名
                data2FileName = actualFileName + '_' + year + '-' + month + '-' + day + '.csv'
                # 判断当前名称的文件是否存在
                if os.path.exists(data2FileName):
                    if os.path.getsize(data2FileName):
                        print('已读取' + data2FileName)
                    else:
                        print(data2FileName + '文件存在但为空')
                        continue
                else:
                    # print(data2FileName + '文件不存在')
                    continue
                # 用暂存变量dataTemp读取生成文档名指向的csv文件
                dataTemp = pd.read_csv(data2FileName,
                                       delimiter=";",       # 在国内使用需手动切换国内格式
                                       encoding='utf-8', header=None, skiprows=1, usecols=[0, 1])
                # 去除数据中的. 防止数据被识别为小数

                # 准备参数
                dataTempList = []
                a = 0
                h = 0
                num = [0.0, 0.0, 0.0, 0.0, 0.0]

                # 调试用的信息
                # print('表格模板：')
                # print(dataTempList)
                # print('选取项：')
                # print(dataTemp[0][3])
                # print('所有项：')
                # print(dataTemp)

                # 将文件中的实际值读取出并和平均值一起存入num数组
                for i in range(3, 96):
                    p = dataTemp[1][i]
                    num[(i + 1) % 4] = float(p)  # 读取每小时的4个数据单元格
                    if (i + 1) % 4 == 3:
                        num[4] = num[0] + num[1] + num[2] + num[3]
                        num[4] = num[4] / 4  # 求它们的平均数
                        num[4] = num[4] * 1000  # 数值扩大1000倍
                        a = a + 1
                        dataTempList.append([y, m, d, a, num[4]])  # 将时间数据和每小时平均数添加进dataTempList列表
                dataTempList.append([y, m, d, 24, 0])

                # 将 列表list 转换为 DataFrame格式
                data2TempDataFrame = pd.DataFrame(dataTempList)
                data2TempDataFrame = data2TempDataFrame.fillna(0)      # 用0替换DataFrame对象中所有的空值
                # print('表格雏形：')
                # print(data2TempDataFrame)

                # 将时间信息与每小时平均数一并存入csv文件：
                data2TempDataFrame.to_csv(actualProductionCsvName, mode='a', header=False, index=None)


# data2读取实际数据csv
data2 = pd.read_csv(actualProductionCsvName, encoding='utf-8')
data2 = data2.round(decimals=2)         # 表格数据只保留两位小数
data2.to_excel('somethingYouNeedToo.xlsx', sheet_name='Production', index=None)
data2.to_excel(actualFileName+'-actualProduction.xlsx', sheet_name='Production', index=None)

# 读取需要录入的表格
wb1 = load_workbook('somethingYouNeed.xlsx')
wb3 = load_workbook(writeFileName + '.xlsx')

# 读取workbook中所有表格
print('请忽略以下4条报警：')
sheets1 = wb1.get_sheet_names()
sheets3 = wb3.get_sheet_names()

# 遍历每个sheet的数据
sheet1 = wb1.get_sheet_by_name(sheets1[0])
sheet3 = wb3.get_sheet_by_name(sheets3[0])

max_row = sheet3.max_row  # 最大行数
max_column = sheet3.max_column  # 最大列数


# 构建数据填写函数
# 用法：
# columnRead：指定读取文档中的具体列标号
# columnWrite：指定写入文档中的具体列标号
# rowWriteStart：开始写入的行标号
# rowWriteStop：停止写入的行标号
# rowReadFixed：读取与写入行之间的差值
def xlsxDataCopy(columnRead, columnWrite, rowWriteStart, rowWriteStop, rowReadFixed):
    "根据行列编号复制单元格内容"
    for m in range(rowWriteStart, rowWriteStop):
        n = columnWrite
        o = columnRead
        n = chr(n)
        o = chr(o)
        i = '%s%d' % (o, m + rowReadFixed)  # 读取单元格编号
        j = '%s%d' % (n, m)  # 写入单元格编号
        cell1 = sheet1[i].value  # 获取data单元格数据
        sheet3[j].value = cell1  # 赋值到test单元格


# 写入数据：
print("写入中。。。")

# 写入预测数据
xlsxDataCopy(108, 101, 2, max_row + 1 - 2160, 2160)
xlsxDataCopy(108, 101, 6602, 8018, -6600)
xlsxDataCopy(108, 101, 8018, 8042, -6624)
xlsxDataCopy(108, 101, 8042, max_row + 1, -6624)

# 写入实际数据


# 完成收尾工作
wb3.save(resultFileName + '.xlsx')  # 保存数据
print("写入完成")
os.remove('somethingYouNeed.xlsx')
os.remove('somethingYouNeedToo.csv')
os.remove('somethingYouNeedToo.xlsx')
wb1.close()  # 关闭excel
wb3.close()


# 一些旧代码：
# 写入预测数据旧代码，现已使用函数代替
# # 填写4-1到12-31数据
# print("写入中。。。")
# for m in range(2, max_row + 1 - 2160):
#     n = 101  # chr(97)='a'
#     o = 108
#     n = chr(n)  # ASCII字符
#     o = chr(o)
#     i = '%s%d' % (o, m + 2160)  # 读取单元格编号
#     j = '%s%d' % (n, m)  # 写入单元格编号
#     cell1 = sheet1[i].value  # 获取data单元格数据
#     sheet2[j].value = cell1  # 赋值到test单元格
#
# # 填写1-1到2-28数据
# for m in range(6602, 8018):
#     n = 101  # chr(97)='a'
#     o = 108
#     n = chr(n)  # ASCII字符
#     o = chr(o)
#     i = '%s%d' % (o, m - 6600)  # 读取单元格编号
#     j = '%s%d' % (n, m)  # 写入单元格编号
#     cell1 = sheet1[i].value  # 获取data单元格数据
#     sheet2[j].value = cell1  # 赋值到test单元格
#
# # 填写2-29数据
# for m in range(8018, 8042):
#     n = 101  # chr(97)='a'
#     o = 108
#     n = chr(n)  # ASCII字符
#     o = chr(o)
#     i = '%s%d' % (o, m - 6624)  # 读取单元格编号
#     j = '%s%d' % (n, m)  # 写入单元格编号
#     cell1 = sheet1[i].value  # 获取data单元格数据
#     sheet2[j].value = cell1  # 赋值到test单元格
#
# # 填写3-1到3-31数据
# for m in range(8042, max_row + 1):
#     n = 101  # chr(97)='a'
#     o = 108
#     n = chr(n)  # ASCII字符
#     o = chr(o)
#     i = '%s%d' % (o, m - 6624)  # 读取单元格编号
#     j = '%s%d' % (n, m)  # 写入单元格编号
#     cell1 = sheet1[i].value  # 获取data单元格数据
#     sheet2[j].value = cell1  # 赋值到test单元格

